import requests
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os
import xlwings as xw

def fetch_crypto_prices():
    url = "https://api.coingecko.com/api/v3/coins/markets"
    params = {
        "vs_currency": "pln",
        "order": "market_cap_desc",
        "per_page": 5,
        "page": 1,
        "sparkline": False,
        "price_change_percentage": "24h"
    }

    response = requests.get(url, params=params)
    data = response.json()

    df = pd.DataFrame(data)
    df = df[["symbol", "name", "current_price", "price_change_percentage_24h", "market_cap", "total_volume"]]

    return df

def update_excel(df):

    try:
        book = xw.Book('crypto.xlsm')
        book.close()
    except Exception as e:
        print(e)
        
    today = datetime.today().strftime('%Y-%m-%d')
    script_dir = os.path.dirname(os.path.realpath(__file__))
    excel_file = os.path.join(script_dir, "crypto.xlsm")
    wb = load_workbook(excel_file, read_only=False, keep_vba=True)
    sheet = wb.active

    sheet.title = today

    headers = ["Symbol", "Name", "Current Price", "Price Change (24h)", "Market Cap", "24h Volume"]
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header)

    for r_idx, row in enumerate(df.values, start=2):
        for c_idx, value in enumerate(row, start=1):
            sheet.cell(row=r_idx, column=c_idx, value=value)

    wb.save(excel_file)
    wb.close()
    
    try:
        book = xw.Book('crypto.xlsm')
        book.open()
    except Exception as e:
        print(e)

if __name__ == "__main__":
    crypto_prices = fetch_crypto_prices()
    update_excel(crypto_prices)
